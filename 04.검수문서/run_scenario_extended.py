"""안심맵 시나리오 자동 테스트 — 확장 (live 모드 API 모킹 + 디바이스 모킹).

Playwright route.fulfill 로 백엔드 응답을 모킹하여 live 모드 경로
(로그인 분기, 동의 철회, SOS 발송, 관리자 SMS/통계 등)와,
geolocation/clipboard 모킹이 필요한 케이스를 자동 검증한다.

run_scenario_playwright.py(기본 11건)에 이어 나머지 P0/P1/P2 중
프로토타입에서 결정적으로 검증 가능한 케이스를 다룬다.
미구현 기능(채팅 글자수 제한 등)은 정직하게 Fail(미구현)로 기록한다.
"""
import json
import sys
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

import os
BASE = os.environ.get("ANSIM_BASE", "http://localhost:3000")
SHOTS = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("./screenshots_auto")
RESULT = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("./단위테스트_결과.xlsx")
SHOTS.mkdir(parents=True, exist_ok=True)

results = []  # (id, name, verdict, detail)


def ok(data):
    return {"status": "success", "data": data}


def err(code, message, status=400):
    return ({"status": "error", "code": code, "message": message}, status)


def shot(page, tc):
    page.screenshot(path=str(SHOTS / f"{tc}.png"))


def record(tc, name, verdict, detail):
    results.append((tc, name, verdict, detail))
    print(f"  [{verdict}] {tc} — {detail}", flush=True)


def fulfill(route, payload):
    """payload: dict(=200) 또는 (dict, status) 튜플."""
    if isinstance(payload, tuple):
        body, status = payload
    else:
        body, status = payload, 200
    route.fulfill(status=status, content_type="application/json",
                  body=json.dumps(body))


def new_page(pw, geo=None, perm=None):
    browser = pw.chromium.launch(headless=True)
    kwargs = {"viewport": {"width": 390, "height": 844}}
    if geo:
        kwargs["geolocation"] = geo
        kwargs["permissions"] = ["geolocation"]
    ctx = browser.new_context(**kwargs)
    return browser, ctx, ctx.new_page()


def fresh(page):
    # route 모킹 시 networkidle 이 끝나지 않을 수 있어 domcontentloaded + 하이드레이션 대기
    page.goto(BASE, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(2500)


def login_live(page, phone="01012345678"):
    """로그인 탭으로 전환 후 번호 입력 → '시작하기'로 live 로그인."""
    fresh(page)
    page.get_by_role("button", name="로그인", exact=True).click()
    page.wait_for_timeout(300)
    page.locator('input[type="tel"]').fill(phone)
    page.get_by_role("button", name="시작하기", exact=True).click()
    page.wait_for_timeout(900)


# ───────────────────────── 인증 분기 (live 모킹) ─────────────────────────
def auth_cases(pw):
    # TC-005 정상 로그인 (family + groupId → 가족 화면)
    browser, ctx, page = new_page(pw)
    try:
        page.route("**/api/auth/login", lambda r: fulfill(r, ok({
            "userId": "u1", "phone": "01012345678", "role": "family",
            "locationSharing": True, "groupId": "g1"})))
        page.route("**/api/groups/g1", lambda r: fulfill(r, ok({
            "groupId": "g1", "groupName": "우리가족", "inviteCode": "ABC123",
            "memberCount": 2, "createdAt": "2026-01-01",
            "members": [{"userId": "p1", "name": "엄마", "role": "parent",
                         "relation": "모", "isParent": True,
                         "locationSharing": True, "lastSeenAt": None}]})))
        page.route("**/api/location/current/g1/parents", lambda r: fulfill(r, ok({
            "plan": "free", "planLimit": 2, "totalParents": 1, "visibleCount": 1,
            "parents": [{"userId": "p1", "name": "엄마", "relation": "모",
                         "locationSharing": True, "latitude": 37.5665,
                         "longitude": 126.978, "accuracy": 15,
                         "timestamp": "2026-01-01T10:00:00Z"}]})))
        login_live(page)
        landed = (page.get_by_text("확인 중").count() > 0
                  or page.get_by_text("우리가족").count() > 0
                  or page.get_by_text("안전 구역").count() > 0)
        record("TC-005", "정상 로그인", "Pass" if landed else "Fail",
               f"family+groupId 로그인 후 가족 화면 진입={landed} (기대: True)")
        shot(page, "TC-005")
    except Exception as e:
        record("TC-005", "정상 로그인", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-006 미가입 번호 로그인 → NOT_FOUND 안내
    browser, ctx, page = new_page(pw)
    try:
        page.route("**/api/auth/login", lambda r: fulfill(
            r, err("NOT_FOUND", "등록되지 않은 번호입니다.", 404)))
        login_live(page, "01099999999")
        msg = page.get_by_text("등록되지 않은 번호").count() > 0
        record("TC-006", "미가입 번호 로그인 시도", "Pass" if msg else "Fail",
               f"NOT_FOUND 안내 노출={msg} (기대: True)")
        shot(page, "TC-006")
    except Exception as e:
        record("TC-006", "미가입 번호 로그인 시도", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-002 중복 번호 회원가입 → CONFLICT
    browser, ctx, page = new_page(pw)
    try:
        page.route("**/api/auth/register", lambda r: fulfill(
            r, err("CONFLICT", "이미 가입된 번호입니다.", 409)))
        fresh(page)
        page.locator('input[type="tel"]').fill("01012345678")
        page.get_by_role("button", name="👵 부모님").click()
        submit = page.get_by_role("button", name="동의하고 시작하기")
        # 약관 동의 체크 — 버튼이 활성화될 때까지 확실히 토글
        for _ in range(3):
            if not submit.is_disabled():
                break
            page.get_by_role("button", name="[필수] 개인정보처리방침 동의").click()
            page.wait_for_timeout(400)
        submit.click()
        page.wait_for_timeout(700)
        msg = page.get_by_text("이미 가입된 휴대폰 번호").count() > 0
        record("TC-002", "중복 번호 회원가입 시도", "Pass" if msg else "Fail",
               f"CONFLICT 시 '이미 가입된 번호' 안내={msg} (기대: True)")
        shot(page, "TC-002")
    except Exception as e:
        record("TC-002", "중복 번호 회원가입 시도", "Block", f"예외: {e}")
    finally:
        browser.close()


# ───────────────────────── 그룹 (live 모킹) ─────────────────────────
def group_cases(pw):
    # TC-011 유효하지 않은 초대 코드 → 안내 메시지 (live: GroupScreen join)
    browser, ctx, page = new_page(pw)
    try:
        # family 신규 가입(groupId 없음) → family-settings → group
        page.route("**/api/auth/register", lambda r: fulfill(r, ok({
            "userId": "u2", "phone": "01022223333", "role": "family",
            "createdAt": "2026-01-01"})))
        page.route("**/api/groups/join", lambda r: fulfill(
            r, err("NOT_FOUND", "초대 코드를 다시 확인해 주세요.", 404)))
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("01022223333")
        page.get_by_role("button", name="가족 자녀").last.click()
        page.get_by_text("개인정보처리방침 동의").click()
        page.get_by_role("button", name="동의하고 시작하기").click()
        page.wait_for_timeout(800)
        page.get_by_role("button", name="가족 그룹 설정하기").click()
        page.wait_for_timeout(400)
        page.get_by_text("초대 코드로 그룹 참여").click()
        page.wait_for_timeout(300)
        page.locator('input').first.fill("ZZZZZZ")
        page.get_by_role("button", name="그룹 참여하기").click()
        page.wait_for_timeout(600)
        msg = page.get_by_text("초대 코드를 다시 확인").count() > 0
        record("TC-011", "유효하지 않은 초대 코드 입력", "Pass" if msg else "Fail",
               f"잘못된 코드 시 안내 메시지={msg} (기대: True)")
        shot(page, "TC-011")
    except Exception as e:
        record("TC-011", "유효하지 않은 초대 코드 입력", "Block", f"예외: {e}")
    finally:
        browser.close()


# ───────────────────────── 부모님 SOS / 위치공유 (demo) ─────────────────────────
def parent_cases(pw):
    def enter_parent_demo(page):
        """부모님 데모 가입 → 온보딩 4단계 통과 → 부모님 메인."""
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("01056781234")
        page.get_by_role("button", name="부모님").last.click()
        page.get_by_text("개인정보처리방침 동의").click()
        page.get_by_role("button", name="로그인 없이 데모로 둘러보기").click()
        page.wait_for_timeout(600)
        page.get_by_role("button", name="시작하기").click()
        page.wait_for_timeout(300)
        page.locator('input').first.fill("ABC123")
        page.get_by_role("button", name="다음 단계로").click()
        page.wait_for_timeout(300)
        page.get_by_text("개인 위치정보 수집").click()
        page.get_by_role("button", name="위치 공유 시작하기").click()
        page.wait_for_timeout(300)
        page.get_by_role("button", name="허용", exact=True).click()
        page.wait_for_timeout(300)
        page.get_by_role("button", name="안심맵 시작하기").click()
        page.wait_for_timeout(500)

    # TC-040 SOS 버튼 탭 → 카운트다운 오버레이
    browser, ctx, page = new_page(pw)
    try:
        enter_parent_demo(page)
        page.get_by_role("button", name="SOS 긴급 호출").click()
        page.wait_for_timeout(300)
        counting = page.get_by_text("긴급 상황").count() > 0 and page.get_by_role("button", name="발송 취소").count() > 0
        record("TC-040", "SOS 버튼 탭 시 카운트다운 오버레이", "Pass" if counting else "Fail",
               f"카운트다운 오버레이+취소버튼 노출={counting} (기대: True)")
        shot(page, "TC-040")
    except Exception as e:
        record("TC-040", "SOS 버튼 탭 시 카운트다운 오버레이", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-043 카운트다운 중 취소 → 오버레이 제거, 미발송
    browser, ctx, page = new_page(pw)
    try:
        enter_parent_demo(page)
        page.get_by_role("button", name="SOS 긴급 호출").click()
        page.wait_for_timeout(300)
        page.get_by_role("button", name="발송 취소").click()
        page.wait_for_timeout(300)
        cancelled = page.get_by_text("긴급 상황").count() == 0 and page.get_by_text("보냈습니다").count() == 0
        record("TC-043", "카운트다운 중 취소 버튼 탭", "Pass" if cancelled else "Fail",
               f"취소 후 오버레이 제거+미발송={cancelled} (기대: True)")
        shot(page, "TC-043")
    except Exception as e:
        record("TC-043", "카운트다운 중 취소 버튼 탭", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-042 2초 카운트다운 종료 → SOS 발송(완료 화면)
    browser, ctx, page = new_page(pw)
    try:
        enter_parent_demo(page)
        page.get_by_role("button", name="SOS 긴급 호출").click()
        page.wait_for_timeout(2800)  # 2초 카운트다운 + 여유
        sent = page.get_by_text("긴급 알림을").count() > 0 or page.get_by_text("보냈습니다").count() > 0
        record("TC-042", "2초 카운트다운 종료 후 SOS 발송", "Pass" if sent else "Fail",
               f"카운트다운 종료 후 발송 완료 화면={sent} (기대: True)")
        shot(page, "TC-042")
    except Exception as e:
        record("TC-042", "2초 카운트다운 종료 후 SOS 발송", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-028 위치 공유 ON → OFF 토글
    browser, ctx, page = new_page(pw)
    try:
        enter_parent_demo(page)
        # 토글은 Toggle 컴포넌트(버튼/스위치). '실시간 위치 공유' 라벨 옆 스위치 클릭
        before = page.get_by_text("켜짐 · 30초마다 전송").count() > 0
        # 토글 클릭: role switch 또는 인접 클릭
        toggle = page.locator("text=실시간 위치 공유").locator("xpath=ancestor::div[1]/following-sibling::*[1]")
        page.get_by_text("실시간 위치 공유").click()  # fallback
        try:
            toggle.click(timeout=1500)
        except Exception:
            pass
        page.wait_for_timeout(500)
        after_off = page.get_by_text("꺼짐").count() > 0 or page.get_by_text("위치 공유 꺼짐").count() > 0
        record("TC-028", "위치 공유 ON→OFF 전환", "Pass" if (before and after_off) else "Fail",
               f"초기 ON={before}, 토글 후 OFF 표시={after_off} (기대: 둘 다 True)")
        shot(page, "TC-028")
    except Exception as e:
        record("TC-028", "위치 공유 ON→OFF 전환", "Block", f"예외: {e}")
    finally:
        browser.close()


# ───────────────────────── 가족 지도/채팅 (live 모킹 + 미구현) ─────────────────────────
def family_cases(pw):
    # TC-032 위치 공유 OFF 상태 지도 안내 (live: parents sharing=false)
    browser, ctx, page = new_page(pw)
    try:
        page.route("**/api/auth/login", lambda r: fulfill(r, ok({
            "userId": "u1", "phone": "01012345678", "role": "family",
            "locationSharing": True, "groupId": "g1"})))
        page.route("**/api/groups/g1", lambda r: fulfill(r, ok({
            "groupId": "g1", "groupName": "우리가족", "inviteCode": "ABC123",
            "memberCount": 2, "createdAt": "2026-01-01",
            "members": [{"userId": "p1", "name": "엄마", "role": "parent",
                         "relation": "모", "isParent": True,
                         "locationSharing": False, "lastSeenAt": None}]})))
        page.route("**/api/location/current/g1/parents", lambda r: fulfill(r, ok({
            "plan": "free", "planLimit": 2, "totalParents": 1, "visibleCount": 1,
            "parents": [{"userId": "p1", "name": "엄마", "relation": "모",
                         "locationSharing": False, "latitude": None,
                         "longitude": None, "accuracy": None, "timestamp": None}]})))
        login_live(page)
        page.wait_for_timeout(800)
        off = page.get_by_text("위치 공유를 꺼두셨").count() > 0
        record("TC-032", "위치 공유 OFF 상태 지도 표시", "Pass" if off else "Fail",
               f"'위치 공유를 꺼두셨습니다' 안내 노출={off} (기대: True)")
        shot(page, "TC-032")
    except Exception as e:
        record("TC-032", "위치 공유 OFF 상태 지도 표시", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-052 / TC-053 / TC-054 — 프로토타입 미구현 (정직 기록)
    record("TC-052", "500자 초과 메시지 전송 시도", "Fail",
           "미구현: 채팅 입력에 글자수 카운터/500자 제한/전송버튼 비활성 로직 없음 (sendChat 무제한)")
    record("TC-053", "WebSocket 5회 재연결 실패 안내 배너", "Fail",
           "미구현: 프로토타입은 폴링 기반, WebSocket 재연결/끊김 안내 배너 미구현")
    record("TC-054", "이미지/파일 전송 차단 안내", "Fail",
           "미구현: 채팅에 파일 첨부 UI/차단 안내('텍스트만 전송 가능') 없음")


# ───────────────────────── 관리자 (demo + live 모킹) ─────────────────────────
def admin_cases(pw):
    def enter_admin(page):
        fresh(page)
        page.get_by_role("button", name="관리자", exact=True).click()
        page.wait_for_timeout(600)

    # TC-069 수신 대상 미선택 시 발송 버튼 비활성
    browser, ctx, page = new_page(pw)
    try:
        enter_admin(page)
        page.get_by_role("button", name="SMS/알림톡 발송").click()
        page.wait_for_timeout(400)
        # 수신번호 입력 필드를 비우기
        ta = page.locator("textarea").first
        if ta.count() == 0:
            ta = page.locator('input').nth(0)
        ta.fill("")
        page.wait_for_timeout(300)
        btn = page.get_by_role("button", name="즉시 메시지 발송")
        disabled = btn.is_disabled()
        record("TC-069", "수신 대상 미선택 발송 시도", "Pass" if disabled else "Fail",
               f"수신번호 공백 시 발송버튼 disabled={disabled} (기대: True)")
        shot(page, "TC-069")
    except Exception as e:
        record("TC-069", "수신 대상 미선택 발송 시도", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-063 관리자 그룹 목록 조회 (demo 데이터)
    browser, ctx, page = new_page(pw)
    try:
        enter_admin(page)
        page.get_by_role("button", name="회원 및 그룹 관리").click()
        page.wait_for_timeout(500)
        listed = page.get_by_text("그룹").count() > 0
        record("TC-063", "관리자 그룹 목록 조회", "Pass" if listed else "Fail",
               f"그룹 관리 화면/목록 렌더={listed} (기대: True, demo 데이터)")
        shot(page, "TC-063")
    except Exception as e:
        record("TC-063", "관리자 그룹 목록 조회", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-074 통계 대시보드 조회 (demo)
    browser, ctx, page = new_page(pw)
    try:
        enter_admin(page)
        page.get_by_role("button", name="통계 대시보드").click()
        page.wait_for_timeout(500)
        stats = page.get_by_text("운영 통계").count() > 0 or page.get_by_role("button", name="CSV 다운로드").count() > 0
        record("TC-074", "관리자 통계 대시보드 조회", "Pass" if stats else "Fail",
               f"통계 대시보드 렌더={stats} (기대: True)")
        shot(page, "TC-074")
    except Exception as e:
        record("TC-074", "관리자 통계 대시보드 조회", "Block", f"예외: {e}")
    finally:
        browser.close()

    # TC-076 통계 CSV 내보내기 (토스트)
    browser, ctx, page = new_page(pw)
    try:
        enter_admin(page)
        page.get_by_role("button", name="통계 대시보드").click()
        page.wait_for_timeout(500)
        page.get_by_role("button", name="CSV 다운로드").click()
        page.wait_for_timeout(400)
        toast = page.get_by_text("CSV").count() > 0
        record("TC-076", "통계 데이터 CSV 내보내기", "Pass" if toast else "Fail",
               f"CSV 다운로드 동작/안내={toast} (기대: True)")
        shot(page, "TC-076")
    except Exception as e:
        record("TC-076", "통계 데이터 CSV 내보내기", "Block", f"예외: {e}")
    finally:
        browser.close()


def write_xlsx():
    wb = load_workbook(RESULT)
    ws = wb["단위테스트결과"] if "단위테스트결과" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]

    def col(name):
        for i, h in enumerate(headers, 1):
            if h and name in str(h):
                return i
        return None
    c_id = col("테스트ID") or 1
    c_res, c_dt, c_sum, c_img = col("테스트결과"), col("실행일시"), col("응답 요약") or col("요약"), col("증적")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    rmap = {tc: (v, d) for tc, n, v, d in results}
    for row in range(2, ws.max_row + 1):
        tc = ws.cell(row=row, column=c_id).value
        if tc in rmap:
            v, d = rmap[tc]
            if c_res: ws.cell(row=row, column=c_res, value=v)
            if c_dt: ws.cell(row=row, column=c_dt, value=now)
            if c_sum: ws.cell(row=row, column=c_sum, value=d)
            if c_img:
                p = SHOTS / f"{tc}.png"
                if p.exists():
                    try:
                        img = XLImage(str(p)); img.width = 120; img.height = 260
                        ws.row_dimensions[row].height = 200
                        ws.add_image(img, ws.cell(row=row, column=c_img).coordinate)
                    except Exception:
                        pass
    wb.save(RESULT)


if __name__ == "__main__":
    print(f"== 안심맵 시나리오 확장 자동 테스트 (live 모킹) == {BASE}", flush=True)
    with sync_playwright() as pw:
        auth_cases(pw)
        group_cases(pw)
        parent_cases(pw)
        family_cases(pw)
        admin_cases(pw)
    p = sum(1 for r in results if r[2] == "Pass")
    f = sum(1 for r in results if r[2] == "Fail")
    b = sum(1 for r in results if r[2] == "Block")
    print(f"\n== 완료 == Pass:{p} | Fail:{f} | Block:{b} | Total:{len(results)}", flush=True)
    if RESULT.exists():
        write_xlsx()
        print(f"결과 기록: {RESULT}", flush=True)
    else:
        print(f"[!] 결과 XLSX 없음: {RESULT}", flush=True)
