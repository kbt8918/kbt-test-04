"""테스트시나리오.md → 테스트시나리오.xlsx 변환

auto-scen-test 스킬 필드 매핑:
  TC-ID         → id
  기능 ID       → scenario
  테스트명      → 테스트명 (보조)
  입력 데이터   → query
  기대 결과     → checkpoint
"""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = Path(__file__).parent / "테스트시나리오.md"
DST = Path(__file__).parent / "테스트시나리오.xlsx"

# ── 파서 ──────────────────────────────────────────────────────────────
TC_HEAD  = re.compile(r"^###\s+(TC-\d+)\.\s+(.+)$")
ROW_LINE = re.compile(r"^\|\s*(.+?)\s*\|\s*(.+?)\s*\|$")

cases = []
current = {}

with open(SRC, encoding="utf-8") as f:
    for line in f:
        line = line.rstrip()
        m = TC_HEAD.match(line)
        if m:
            if current.get("id"):
                cases.append(current)
            current = {"id": m.group(1), "테스트명": m.group(2)}
            continue
        m = ROW_LINE.match(line)
        if m:
            key, val = m.group(1).strip(), m.group(2).strip()
            if key == "관련 기능":
                current["scenario"] = val
            elif key == "입력 데이터":
                current["query"] = val
            elif key == "기대 결과":
                current["checkpoint"] = val
            elif key == "사전 조건":
                current["사전조건"] = val
            elif key == "테스트 단계":
                current["테스트단계"] = val

if current.get("id"):
    cases.append(current)

print(f"케이스 파싱: {len(cases)}건")

# ── XLSX 생성 ─────────────────────────────────────────────────────────
COLS = ["id", "scenario", "테스트명", "사전조건", "테스트단계", "query", "checkpoint"]

wb = Workbook()
ws = wb.active
ws.title = "시나리오"
ws.append(COLS)

for c in cases:
    ws.append([c.get(col, "") for col in COLS])

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

WIDTHS = {"id": 12, "scenario": 10, "테스트명": 30, "사전조건": 40,
          "테스트단계": 50, "query": 40, "checkpoint": 60}
for i, col in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(col, 18)
ws.row_dimensions[1].height = 30

wb.save(DST)
print(f"저장: {DST}")
