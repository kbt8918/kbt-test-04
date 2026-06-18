"""안심맵 시나리오 자동 테스트 (webapp-testing / Playwright 직접 제어).

프로토타입의 데모 모드(백엔드 비의존, 결정적 UI 동작)를 구동해
테스트시나리오.md의 P0 검증 가능 케이스를 자동 검증한다.
각 케이스: 동작 수행 -> 기대 결과 단언 -> 스크린샷 -> Pass/Fail 기록.
"""
import sys
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

BASE = "http://localhost:3000"
SHOTS = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("./screenshots_auto")
RESULT = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("./단위테스트_결과.xlsx")
SHOTS.mkdir(parents=True, exist_ok=True)

results = []  # (id, name, verdict, detail)


def shot(page, tc):
    p = SHOTS / f"{tc}.png"
    page.screenshot(path=str(p))
    return p


def record(tc, name, verdict, detail):
    results.append((tc, name, verdict, detail))
    print(f"  [{verdict}] {tc} — {detail}", flush=True)


def fresh(page):
    """홈(로그인/회원가입 탭)으로 초기화."""
    page.goto(BASE, wait_until="networkidle")
    page.wait_for_timeout(400)


def to_group_screen(page, phone):
    """가족 자녀 데모로 가입 -> family-settings -> '가족 그룹 설정하기' -> GroupScreen."""
    fresh(page)
    page.get_by_role("button", name="회원가입", exact=True).click()
    page.locator('input[type="tel"]').fill(phone)
    page.get_by_role("button", name="가족 자녀").last.click()
    page.get_by_text("개인정보처리방침 동의").click()
    page.get_by_role("button", name="로그인 없이 데모로 둘러보기").click()
    page.wait_for_timeout(700)
    page.get_by_role("button", name="가족 그룹 설정하기").click()
    page.wait_for_timeout(500)


def run(pw):
    browser = pw.chromium.launch(headless=True)
    ctx = browser.new_context(viewport={"width": 390, "height": 844})
    page = ctx.new_page()

    # ── TC-004: 약관 미동의 시 가입 버튼 비활성화 ──
    try:
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("01012345678")
        page.wait_for_timeout(200)
        btn = page.get_by_role("button", name="동의하고 시작하기")
        disabled = btn.is_disabled()
        record("TC-004", "약관 미동의 가입 시도",
               "Pass" if disabled else "Fail",
               f"동의 체크 전 가입버튼 disabled={disabled} (기대: True)")
        shot(page, "TC-004")
    except Exception as e:
        record("TC-004", "약관 미동의 가입 시도", "Block", f"예외: {e}")

    # ── TC-001: 정상 회원가입 (약관 동의 후 버튼 활성 + 부모님 데모 진입) ──
    try:
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("01012345678")
        # 역할 부모님 선택
        page.get_by_role("button", name="부모님").last.click()
        page.get_by_text("개인정보처리방침 동의").click()
        page.wait_for_timeout(200)
        btn = page.get_by_role("button", name="동의하고 시작하기")
        enabled = not btn.is_disabled()
        # 데모로 둘러보기 -> 부모님 역할이면 온보딩 진입
        page.get_by_role("button", name="로그인 없이 데모로 둘러보기").click()
        page.wait_for_timeout(600)
        onboarding = page.get_by_text("안심맵 온보딩").count() > 0
        ok = enabled and onboarding
        record("TC-001", "정상 회원가입",
               "Pass" if ok else "Fail",
               f"동의후 버튼활성={enabled}, 가입후 온보딩진입={onboarding} (기대: 둘 다 True)")
        shot(page, "TC-001")
    except Exception as e:
        record("TC-001", "정상 회원가입", "Block", f"예외: {e}")

    # ── TC-003: 잘못된 번호(9자리) 입력 시 버튼 비활성 ──
    try:
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("010123456")  # 9 digits
        page.get_by_text("개인정보처리방침 동의").click()
        page.wait_for_timeout(200)
        btn = page.get_by_role("button", name="동의하고 시작하기")
        disabled = btn.is_disabled()
        record("TC-003", "잘못된 번호 형식 입력",
               "Pass" if disabled else "Fail",
               f"9자리 입력 시 가입버튼 disabled={disabled} (기대: True, 10자리 미만 무효)")
        shot(page, "TC-003")
    except Exception as e:
        record("TC-003", "잘못된 번호 형식 입력", "Block", f"예외: {e}")

    # ── TC-021: 개인정보처리방침 미동의 시 가입 버튼 비활성 (TC-004와 동일 게이팅, 별도 단언) ──
    try:
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("01055667788")
        page.wait_for_timeout(200)
        before = page.get_by_role("button", name="동의하고 시작하기").is_disabled()
        page.get_by_text("개인정보처리방침 동의").click()
        page.wait_for_timeout(200)
        after = page.get_by_role("button", name="동의하고 시작하기").is_disabled()
        ok = before and not after
        record("TC-021", "개인정보처리방침 미동의 시 버튼 비활성",
               "Pass" if ok else "Fail",
               f"동의전 disabled={before}, 동의후 disabled={after} (기대: True->False)")
        shot(page, "TC-021")
    except Exception as e:
        record("TC-021", "개인정보처리방침 미동의 시 버튼 비활성", "Block", f"예외: {e}")

    # ── TC-079: 개인정보처리방침 '보기' -> 토스트 안내 ──
    try:
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.get_by_role("button", name="보기", exact=True).click()
        page.wait_for_timeout(400)
        toast = page.get_by_text("개인정보처리방침 페이지").count() > 0
        record("TC-079", "개인정보처리방침 페이지 접근",
               "Pass" if toast else "Fail",
               f"'보기' 탭 시 개인정보처리방침 안내 노출={toast} (기대: True)")
        shot(page, "TC-079")
    except Exception as e:
        record("TC-079", "개인정보처리방침 페이지 접근", "Block", f"예외: {e}")

    # ── TC-008: 정상 가족 그룹 생성 (가족 데모 -> 그룹생성 -> 초대코드/QR) ──
    try:
        to_group_screen(page, "01099887766")
        # 그룹 선택 화면 -> 새 그룹 생성
        page.get_by_text("새로운 가족 그룹 생성").click()
        page.wait_for_timeout(300)
        page.locator('input').first.fill("우리 가족")
        page.get_by_role("button", name="그룹 만들기").click()
        page.wait_for_timeout(500)
        done = page.get_by_text("그룹 생성이 완료되었습니다").count() > 0
        code_label = page.get_by_text("초대 코드 (6자리)").count() > 0
        ok = done and code_label
        record("TC-008", "정상 가족 그룹 생성",
               "Pass" if ok else "Fail",
               f"완료화면={done}, 초대코드표시={code_label} (기대: 둘 다 True)")
        shot(page, "TC-008")
    except Exception as e:
        record("TC-008", "정상 가족 그룹 생성", "Block", f"예외: {e}")

    # ── TC-009: 그룹명 미입력 시 생성 버튼 비활성 ──
    try:
        to_group_screen(page, "01099887755")
        page.get_by_text("새로운 가족 그룹 생성").click()
        page.wait_for_timeout(300)
        disabled = page.get_by_role("button", name="그룹 만들기").is_disabled()
        record("TC-009", "그룹명 미입력 생성 시도",
               "Pass" if disabled else "Fail",
               f"그룹명 공백 시 생성버튼 disabled={disabled} (기대: True)")
        shot(page, "TC-009")
    except Exception as e:
        record("TC-009", "그룹명 미입력 생성 시도", "Block", f"예외: {e}")

    # ── TC-012: 정원 초과(FULL00) 그룹 참여 시 오류 메시지 ──
    try:
        to_group_screen(page, "01099880011")
        page.get_by_text("초대 코드로 그룹 참여").click()
        page.wait_for_timeout(300)
        page.locator('input').first.fill("FULL00")
        page.get_by_role("button", name="그룹 참여하기").click()
        page.wait_for_timeout(400)
        err = page.get_by_text("한도").count() > 0 or page.get_by_text("초과").count() > 0
        record("TC-012", "5명 정원 초과 그룹 참여 시도",
               "Pass" if err else "Fail",
               f"FULL00 입력 시 정원초과 오류 노출={err} (기대: True)")
        shot(page, "TC-012")
    except Exception as e:
        record("TC-012", "5명 정원 초과 그룹 참여 시도", "Block", f"예외: {e}")

    # ── TC-014: 온보딩 4단계 정상 완료 (부모님 데모) ──
    try:
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("01056781234")
        page.get_by_role("button", name="부모님").last.click()
        page.get_by_text("개인정보처리방침 동의").click()
        page.get_by_role("button", name="로그인 없이 데모로 둘러보기").click()
        page.wait_for_timeout(600)
        # step1 시작하기
        page.get_by_role("button", name="시작하기").click()
        page.wait_for_timeout(300)
        # step2 초대코드
        page.locator('input').first.fill("ABC123")
        page.get_by_role("button", name="다음 단계로").click()
        page.wait_for_timeout(300)
        # step3 필수 동의
        page.get_by_text("개인 위치정보 수집").click()
        page.get_by_role("button", name="위치 공유 시작하기").click()
        page.wait_for_timeout(300)
        # geolocation 바텀시트 허용
        page.get_by_role("button", name="허용", exact=True).click()
        page.wait_for_timeout(400)
        done = page.get_by_text("설정이 완료되었어요").count() > 0
        record("TC-014", "온보딩 4단계 정상 완료",
               "Pass" if done else "Fail",
               f"4단계(설정 완료) 도달={done} (기대: True)")
        shot(page, "TC-014")
    except Exception as e:
        record("TC-014", "온보딩 4단계 정상 완료", "Block", f"예외: {e}")

    # ── TC-019: 위치정보 동의 필수 항목 미체크 시 '위치 공유 시작' 버튼 비활성 ──
    try:
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("01056789999")
        page.get_by_role("button", name="부모님").last.click()
        page.get_by_text("개인정보처리방침 동의").click()
        page.get_by_role("button", name="로그인 없이 데모로 둘러보기").click()
        page.wait_for_timeout(600)
        page.get_by_role("button", name="시작하기").click()
        page.wait_for_timeout(300)
        page.locator('input').first.fill("ABC123")
        page.get_by_role("button", name="다음 단계로").click()
        page.wait_for_timeout(300)
        # 필수 동의 체크 안 함
        disabled = page.get_by_role("button", name="위치 공유 시작하기").is_disabled()
        record("TC-019", "필수 항목 미동의 시 확인버튼 비활성",
               "Pass" if disabled else "Fail",
               f"위치정보 필수 미동의 시 버튼 disabled={disabled} (기대: True)")
        shot(page, "TC-019")
    except Exception as e:
        record("TC-019", "필수 항목 미동의 시 확인버튼 비활성", "Block", f"예외: {e}")

    # ── TC-011-2: 데모/미가입 상태에서 초대 링크 복사 차단 ──
    try:
        fresh(page)
        page.get_by_role("button", name="회원가입", exact=True).click()
        page.locator('input[type="tel"]').fill("01077665544")
        page.get_by_role("button", name="가족 자녀").last.click()
        page.get_by_text("개인정보처리방침 동의").click()
        page.get_by_role("button", name="로그인 없이 데모로 둘러보기").click()
        page.wait_for_timeout(700)
        # family-settings 의 '가족 구성원 초대하기' 탭 -> 데모 차단 토스트
        page.get_by_text("가족 구성원 초대하기").click()
        page.wait_for_timeout(400)
        blocked = (page.get_by_text("데모 체험 중").count() > 0
                   or page.get_by_text("먼저 생성").count() > 0)
        record("TC-011-2", "데모/미가입 초대 링크 복사 차단",
               "Pass" if blocked else "Fail",
               f"복사 차단 토스트 노출={blocked} (기대: True, 데모/미가입 시 복사 중단)")
        shot(page, "TC-011-2")
    except Exception as e:
        record("TC-011-2", "데모/미가입 초대 링크 복사 차단", "Block", f"예외: {e}")

    browser.close()


def write_xlsx():
    wb = load_workbook(RESULT)
    ws = wb["단위테스트결과"] if "단위테스트결과" in wb.sheetnames else wb.active
    # 헤더 행 찾기
    header_row = 1
    headers = [c.value for c in ws[header_row]]
    def col(name):
        for i, h in enumerate(headers, 1):
            if h and name in str(h):
                return i
        return None
    c_id = col("테스트ID") or col("ID") or 1
    c_res = col("테스트결과")
    c_dt = col("실행일시")
    c_sum = col("응답 요약") or col("요약")
    c_img = col("증적")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    rmap = {tc: (v, d, n) for tc, n, v, d in results}
    for row in range(header_row + 1, ws.max_row + 1):
        tc = ws.cell(row=row, column=c_id).value
        if tc in rmap:
            v, d, n = rmap[tc]
            if c_res: ws.cell(row=row, column=c_res, value=v)
            if c_dt: ws.cell(row=row, column=c_dt, value=now)
            if c_sum: ws.cell(row=row, column=c_sum, value=d)
            if c_img:
                p = SHOTS / f"{tc}.png"
                if p.exists():
                    try:
                        img = XLImage(str(p)); img.width = 120; img.height = 260
                        ws.row_dimensions[row].height = 200
                        ws.add_image(img, ws.cell(row=row, column=c_img).coordinate)
                    except Exception:
                        pass
    wb.save(RESULT)


if __name__ == "__main__":
    print(f"== 안심맵 시나리오 자동 테스트 (demo 모드) == {BASE}", flush=True)
    with sync_playwright() as pw:
        run(pw)
    p = sum(1 for r in results if r[2] == "Pass")
    f = sum(1 for r in results if r[2] == "Fail")
    b = sum(1 for r in results if r[2] == "Block")
    print(f"\n== 완료 == Pass:{p} | Fail:{f} | Block:{b} | Total:{len(results)}", flush=True)
    if RESULT.exists():
        write_xlsx()
        print(f"결과 기록: {RESULT}", flush=True)
    else:
        print(f"[!] 결과 XLSX 없음 — build_result_xlsx.py 로 먼저 생성 필요: {RESULT}", flush=True)
