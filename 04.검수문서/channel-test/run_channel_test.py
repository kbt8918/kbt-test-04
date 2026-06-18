"""진입 채널별 접속 테스트 — 안심맵 프로토타입.

각 진입 채널(로그인·온보딩 / 부모님 / 가족 / 관리자 / 초대링크)에 실제 접속하여
해당 화면이 렌더되는지 상단 라우트 라벨로 검증하고 스크린샷을 남긴다.
"""
import sys, datetime
from pathlib import Path
from playwright.sync_api import sync_playwright

BASE = "http://localhost:3002"
OUT = Path(__file__).parent
SHOTS = OUT / "shots"
SHOTS.mkdir(parents=True, exist_ok=True)

# (채널ID, 이름, 진입 동작, 기대 라우트 라벨 일부)
CHANNELS = [
    ("CH-01", "로그인·온보딩", {"goto": BASE, "click": "로그인·온보딩"}, "SCR-001 로그인/회원가입"),
    ("CH-02", "부모님",         {"goto": BASE, "click": "부모님"},       "SCR-003 부모님 메인 (SOS)"),
    ("CH-03", "가족",           {"goto": BASE, "click": "가족"},         "SCR-006 실시간 지도"),
    ("CH-04", "관리자",         {"goto": BASE, "click": "관리자"},       "SCR-011~013 관리자 어드민"),
    ("CH-05", "초대 링크(/join?code=)", {"goto": BASE + "/join?code=ABC123"}, "SCR-001 로그인/회원가입"),
]

results = []

with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    page = browser.new_page(viewport={"width": 1280, "height": 900})
    for cid, name, act, expect in CHANNELS:
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        verdict, note = "Fail", ""
        try:
            page.goto(act["goto"], wait_until="domcontentloaded", timeout=25000)
            page.wait_for_timeout(3500)  # React 하이드레이션 대기
            if "click" in act:
                # 플로우 스위처 탭만 정확히 매칭(화면 내 동명 버튼 회피)
                page.get_by_role("button", name=act["click"], exact=True).first.click(timeout=8000)
            # 상단 라우트 라벨이 기대 화면으로 바뀌는지 확인
            page.wait_for_timeout(1200)
            page.wait_for_selector(f"text={expect}", timeout=8000)
            verdict = "Pass"
            note = f"라우트 라벨 '{expect}' 확인됨"
        except Exception as e:
            note = f"검증 실패: {type(e).__name__}: {str(e)[:120]}"
        shot = SHOTS / f"{cid}.png"
        try:
            page.screenshot(path=str(shot), full_page=True)
        except Exception:
            pass
        results.append((cid, name, act["goto"], expect, verdict, ts, note))
        print(f"{cid} {name}: {verdict} | {note}")
    browser.close()

# 결과 XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

FONT = "맑은 고딕"
HF = PatternFill("solid", start_color="2F5496"); HFONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BFONT = Font(name=FONT, size=10)
thin = Side(style="thin", color="BFBFBF"); BD = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top"); CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
PASS_FILL = PatternFill("solid", start_color="C6EFCE"); FAIL_FILL = PatternFill("solid", start_color="FFC7CE")

wb = Workbook(); ws = wb.active; ws.title = "진입채널 테스트결과"
ws.sheet_view.showGridLines = False
cols = ["채널 ID", "진입 채널", "접속 경로", "기대 화면", "테스트 결과", "실행일시", "비고", "증적"]
for ci, h in enumerate(cols, 1):
    c = ws.cell(1, ci, h); c.fill = HF; c.font = HFONT; c.alignment = CEN; c.border = BD
for ri, (cid, name, url, expect, verdict, ts, note) in enumerate(results, 2):
    vals = [cid, name, url, expect, verdict, ts, note, ""]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v); c.font = BFONT; c.border = BD
        c.alignment = CEN if ci in (1, 5, 6) else WRAP
        if ci == 5:
            c.fill = PASS_FILL if v == "Pass" else FAIL_FILL
            c.font = Font(name=FONT, bold=True, size=10)
    img_path = SHOTS / f"{cid}.png"
    if img_path.exists():
        try:
            img = XLImage(str(img_path)); img.width = 240; img.height = 168
            ws.add_image(img, f"H{ri}")
            ws.row_dimensions[ri].height = 130
        except Exception:
            pass
widths = [10, 20, 30, 26, 12, 19, 40, 36]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
ws.freeze_panes = "A2"

# 요약
ws2 = wb.create_sheet("결과 요약"); ws2.sheet_view.showGridLines = False
ws2["A1"] = "진입 채널별 접속 테스트 — 결과 요약"
ws2["A1"].font = Font(name=FONT, bold=True, size=14, color="2F5496")
npass = sum(1 for r in results if r[4] == "Pass")
nfail = len(results) - npass
rows = [("총 채널 수", len(results)), ("Pass", npass), ("Fail", nfail),
        ("Pass 비율", f"{npass/len(results)*100:.1f}%"),
        ("대상", BASE), ("실행 도구", "Playwright (headless Chromium)")]
r = 3
for k, v in rows:
    ws2.cell(r, 1, k).font = Font(name=FONT, bold=True, size=10)
    ws2.cell(r, 1).fill = PatternFill("solid", start_color="D9E1F2")
    ws2.cell(r, 2, v).font = BFONT
    for ci in (1, 2):
        ws2.cell(r, ci).border = BD; ws2.cell(r, ci).alignment = WRAP
    r += 1
ws2.column_dimensions["A"].width = 16; ws2.column_dimensions["B"].width = 40

out = OUT / "진입채널_테스트결과.xlsx"
wb.save(out)
print(f"\nSAVED {out} | Pass {npass} Fail {nfail} Total {len(results)}")
